from openpyxl import Workbook
from lunarcalendar import Solar, Converter
from datetime import date, timedelta

#lunars = [0x0E, 0x11, 0x12, 0x13]
lunars = [0x0E]

wb = Workbook()
ws = wb.active
ws.cell(1, 1).value = "Gregorian"
for i, l in enumerate(lunars):
    ws.cell(1, 2+i).value = f'Lunar {l:02X}'
ws.cell(1, len(lunars)+2).value = 'Solar2Lunar'

for i in range(1, 100001):
    ws.cell(i+1, 1).value = i
    ws.cell(i+1, 1).number_format = 'yyyy-mm-dd'
    for j, l in enumerate(lunars):
        ws.cell(i+1, j+2).value = f'=TEXT({i}, "[$-{l:02X}0000]yyyy-mm-dd")'

    dt = date(1900, 1, 1) + timedelta(days=i-1)
    try:
        lunar = Converter.Solar2Lunar(Solar(dt.year, dt.month, dt.day))
        ws.cell(i+1, len(lunars)+2).value = f'{lunar.year:04}-{lunar.month:02}-{lunar.day:02}'
    except Exception:
        pass

wb.save(filename=f'lunarcal.xlsx')
