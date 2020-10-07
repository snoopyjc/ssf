from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re

with open('valid.tsv', 'r', encoding='utf-8') as v:
    lines = v.read().splitlines()

CHUNKSIZE=200       # Excel only has room for a small number of custom formats

def divide_chunks(l, n):
    for i in range(0, len(l), n):
        yield l[i:i+n]

data = [12.3456789, -12.3456789, 0, 'abcdef']
for c, chunk in enumerate(divide_chunks(lines, CHUNKSIZE), start=1):
    wb = Workbook()
    ws = wb.active

    ws.cell(1, 1).value = 'Number Format'
    ws.column_dimensions[get_column_letter(1)].width = 48
    for col, d in enumerate(data, start=2):
        ws.cell(1, col).value = d
        ws.column_dimensions[get_column_letter(col)].width = 32

    for row, l in enumerate(chunk, start=2):
        # Remove [ENG] etc which Excel doesn't like
        l = re.sub(r'\[[A-Z][A-Z][A-Z]\]\[', r'[', l)
        ws.cell(row, 1).value = l
        ws.cell(row, 1).number_format = '@'
        for col, d in enumerate(data, start=2):
            ws.cell(row, col).value = d
            ws.cell(row, col).number_format = l

    wb.save(filename=f'valid{c}.xlsx')
   
    
