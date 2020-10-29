from openpyxl import Workbook
#from lunarcalendar import Solar, Converter
from datetime import date, timedelta
from convertdate import islamic
from ummalqura.hijri_date import HijriDate

um = [0x17, 0x06]

wb = Workbook()
ws = wb.active
ws.cell(1, 1).value = "Gregorian"
for i, l in enumerate(um):
    ws.cell(1, 2+i).value = f'Um {l:02X}'
ws.cell(1, len(um)+2).value = 'islamic'

for i in range(1, 13600):
    ws.cell(i+1, 1).value = i
    ws.cell(i+1, 1).number_format = 'yyyy-mm-dd'
    for j, l in enumerate(um):
        ws.cell(i+1, j+2).value = f'=TEXT({i}, "[$-{l:02X}0000]yyyy-mm-dd")'

    dt = date(1900, 1, 1) + timedelta(days=i-1)
    try:
        isl = islamic.from_gregorian(dt.year, dt.month, dt.day)
        ws.cell(i+1, len(um)+2).value = f'{isl[0]:04}-{isl[1]:02}-{isl[2]:02}'
    except Exception as e:
        print(e)
        pass

wb.save(filename=f'umcal.xlsx')
