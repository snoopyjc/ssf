from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re

with open('../ssf/lcid.tsv', 'r', encoding='utf-8') as v:
    lines = v.read().splitlines()

with open('../ssf/numbers.tsv', 'r', encoding='utf-8') as n:
    numbers = n.read().splitlines()

number_names = [ '', 'en', 'hi', 'Hi', 'sa', 'bn',          # 00 - 05
                 'guru', 'gu', 'or', 'ta', 'te',            # 06 - 0A
                 'kn', 'ml', 'th', 'lo', 'bo', 'my',        # 0B - 10
                 'am', 'km', 'mn', '', '',                  # 11 - 15
                 '', '', '', '', '',                        # 16 - 1A
                 'ja', 'jpn', 'ja3', 'zh', 'zhu', 'zhn',    # 1B - 20
                 'zhtl', 'zhtu', 'zhtn', 'ko', 'ko2', 'ko3', 'ko4' ] # 21 - 27

dates = [43836, 43865, 43894, 43923, 43952, 43988, 44017, 44046, 44082, 44111, 44140, 44169, 44198]
desc = ['Mon Jan', 'Tue Feb', 'Wed Mar', 'Thu Apr', 'Fri May', 'Sat Jun', 'Sun Jul', 'Mon Aug', 'Tue Sep', 'Wed Oct', 'Thu Nov', 'Fri Dec', 'Sat Jan']

# For calendars that have leap months, when is the first monday of a leap year?
leap_year_first_monday = {0x08: 44571, 0x0E: 43857+7, 0x11: 43857+7, 0x12: 43857+7, 0x13: 43857+7}
# For 0x0E and friends, we must subtract 364 from our regular dates so they don't fall in a leap year, then add a month
non_leap_year_lunar_offset = -364+35

# Corresponds with calendars we handle in ssf_calendar:
_calendars = [True, True, True, True, True, True,   # 00-05
        True, True, True, True, True,               # 06-0A
        True, True, False,                          # 0B-0D
        True, False, False, True, True, True, False, False, False, 
        True, False, False, False, False, False, False, False, False]

def find(elem, lst):
    try:
        return lst.index(elem)
    except Exception:
        pass
    return -1

for c in range(0x20):
    if not _calendars[c]:
        continue
    for yr in ('', '_leap'):
        if yr and c not in leap_year_first_monday:
            continue
        offset = 0
        if c in leap_year_first_monday:
            lyfm = leap_year_first_monday[c]
            if yr:
                offset = lyfm - dates[0]    # Move us to a leap year
            elif lyfm > dates[0] and lyfm < dates[-1]:
                offset = non_leap_year_lunar_offset    # Move us to a non-leap year
        wb = Workbook()
        ws = wb.active
        ws.cell(1, 1).value = 'LCID'
        ws.cell(1, 2).value = 'Locale'
        for col, d in enumerate(desc, start=3):
            ws.cell(1, col).value = d
            ws.column_dimensions[get_column_letter(col)].width = 38
        for row, line in enumerate(lines[1:], start=2):
            ls = line.split('\t')
            num = int(ls[0], 16)
            lang = ls[1].split('-')[0]
            ndx = find(lang, number_names)
            if ndx > 1:                 # Some of the output is numbers, so localize those too
                num |= ndx << 24
            num |= c << 16          # Insert the calendar
            fmt = f'[$-{num:X}]dddd,ddd,mmmm,mmm,mmmmm,m'
            ws.cell(row, 1).value = f'0x{num:X}'
            ws.cell(row, 2).value = ls[1]
            for col, d in enumerate(dates, start=3):
                #ws.cell(row, col).value = d
                #ws.cell(row, col).number_format = fmt
                ws.cell(row, col).value = f'=TEXT({d+offset}, "{fmt}")'
        wb.save(filename=f'daymonth{c:02X}{yr}.xlsx')
