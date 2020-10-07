from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re

with open('../ssf/lcid.tsv', 'r', encoding='utf-8') as v:
    lines = v.read().splitlines()

LIMIT=12

wb = Workbook()
ws = wb.active
ws.cell(1, 1).value = 'DBNum'
ws.cell(1, 2).value = 'LCID'
ws.cell(1, 3).value = 'Locale'
ws.cell(1, 4).value = 'Number Format'
ws.column_dimensions['D'].width = 38
ws.cell(1, 5).value = '0123456789'
ws.column_dimensions['E'].width = 26
#ws.cell(1, 6).value = '-1.23E+45'
#ws.cell(1, 7).value = '10,000'
for i in range(1, LIMIT+1):
    ws.cell(1, 5+i).value = f'1E{i}'
row = 2
for dbnum in range(1, 4):
    for line in lines[1:]:
        ls = line.split('\t')
        num = int(ls[0], 16)
        if num < 0x404:
            continue
        if 0x405 <= num <= 0x410:
            continue
        if 0x413 <= num <= 0x803:
            continue
        if num >= 0x807:
            continue
        fmt = f'[DBNum{dbnum}][$-{num:X}]0000000000'
        ws.cell(row, 1).value = f'DBNum{dbnum}'
        ws.cell(row, 2).value = f'0x{num:X}'
        ws.cell(row, 3).value = ls[1]
        ws.cell(row, 4).value = fmt + ' (' + ls[1] + ')'
        ws.cell(row, 5).value = 123456789
        ws.cell(row, 5).number_format = fmt
        #fmt1 = f'[DBNum{dbnum}][$-{num:X}]#,###'
        fmt2 = f'[DBNum{dbnum}][$-{num:X}]General'
        #ws.cell(row, 6).value = -1.23E+45
        #ws.cell(row, 6).number_format = fmt2
        #ws.cell(row, 7).value = 10000
        #ws.cell(row, 7).number_format = fmt1
        for i in range(1, LIMIT+1):
            ws.cell(row, 5+i).value = 10**i
            ws.cell(row, 5+i).number_format = fmt2
        row += 1

wb.save(filename='dbnum.xlsx')
