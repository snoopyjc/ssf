from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re

LIMIT=8

wb = Workbook()
ws = wb.active
ws.cell(1, 1).value = 'xx'
ws.cell(1, 2).value = 'Number Format'
ws.column_dimensions['B'].width = 26
ws.cell(1, 3).value = '0123456789'
ws.column_dimensions['C'].width = 30
ws.cell(1, 4).value = '-1.23E+45'
ws.cell(1, 5).value = '-1.23E-45'
ws.cell(1, 6).value = '10,000'
for i in range(1, LIMIT+1):
    ws.cell(1, 6+i).value = f'1E{i}'
row = 2
for xx in range(0, 0x28):
        fmt = f'[$-{xx:X}000000]0000000000'
        ws.cell(row, 1).value = f'0x{xx:X}'
        ws.cell(row, 2).value = fmt
        ws.cell(row, 3).value = 123456789
        ws.cell(row, 3).number_format = fmt
        ws.cell(row, 4).value = -1.23E+45
        ws.cell(row, 4).number_format = f'[$-{xx:X}000000]General'
        ws.cell(row, 5).value = -1.23E-45
        ws.cell(row, 5).number_format = f'[$-{xx:X}000000]General'
        ws.cell(row, 6).value = 10000
        ws.cell(row, 6).number_format = f'[$-{xx:X}000000]#,###'
        fmt2 = f'[$-{xx:X}000000]General'
        for i in range(1, LIMIT+1):
            ws.cell(row, 6+i).value = 10**i
            ws.cell(row, 6+i).number_format = fmt2
        row += 1

wb.save(filename='numbers.xlsx')
