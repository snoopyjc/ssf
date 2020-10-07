from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re

with open('../ssf/lcid.tsv', 'r', encoding='utf-8') as v:
    lines = v.read().splitlines()

CHUNKSIZE=200       # Excel only has room for a small number of custom formats

def divide_chunks(l, n):
    for i in range(0, len(l), n):
        yield l[i:i+n]

for c, chunk in enumerate(divide_chunks(lines[1:], CHUNKSIZE), start=1):
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1).value = 'LCID'
    ws.cell(1, 2).value = 'Locale'
    ws.cell(1, 3).value = '1 AM'
    ws.cell(1, 4).value = '1 PM'
    #ws.cell(1, 5).value = '1 A'
    #ws.cell(1, 6).value = '1 P'
    #ws.cell(1, 7).value = '1 a'
    #ws.cell(1, 8).value = '1 p'
    for row, line in enumerate(chunk, start=2):
        ls = line.split('\t')
        num = int(ls[0], 16)
        #if num < 0x404:
            #continue
        #if 0x405 <= num <= 0x410:
            #continue
        #if 0x413 <= num <= 0x803:
            #continue
        #if num >= 0x807:
            #continue
        fmt1 = f'[$-{num:X}] h AM/PM'
        fmt2 = f'[$-{num:X}] h A/P'
        fmt3 = f'[$-{num:X}] h a/p'
        ws.cell(row, 1).value = f'0x{num:X}'
        ws.cell(row, 2).value = ls[1]
        ws.cell(row, 3).value = 0.0416666666666667  # 1AM
        ws.cell(row, 3).number_format = fmt1
        ws.cell(row, 4).value = 0.541666666666667   # 1PM
        ws.cell(row, 4).number_format = fmt1
        #ws.cell(row, 5).value = 0.0416666666666667  # 1AM
        #ws.cell(row, 5).number_format = fmt2
        #ws.cell(row, 6).value = 0.541666666666667   # 1PM
        #ws.cell(row, 6).number_format = fmt2
        #ws.cell(row, 7).value = 0.0416666666666667  # 1AM
        #ws.cell(row, 7).number_format = fmt3
        #ws.cell(row, 8).value = 0.541666666666667   # 1PM
        #ws.cell(row, 8).number_format = fmt3
    
    wb.save(filename=f'ampm{c}.xlsx')
